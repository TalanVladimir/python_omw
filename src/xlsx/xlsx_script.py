import pandas as pd
import openpyxl
import xlwings as xw

workFile = './src/xlsx/test.xlsx'

# dataFrame = pd.read_excel('test.xlsx')

dataframe = openpyxl.load_workbook(workFile)
 
# Define variable to read sheet
dataframe1 = dataframe.active
 
# Iterate the loop to read the cell values
for row in range(0, dataframe1.max_row):
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        print(f"row: {row} col: {col} value: {col[row].value}")

wb = openpyxl.load_workbook(workFile)

ws = wb['Sheet1']  # or wb.active

ws['E6'] = "value new"

wb.save(workFile)